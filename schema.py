import pandas as pd 
import re 

import ascii_magic
output = ascii_magic.from_image_file("C5-Logo (1).jpg",column=1000,char='#')
ascii_magic.to_terminal(output)

def open_file(file_path):
    if '.xlsx' in file_path:
        df = pd.read_excel(file_path)
        
        l0 = list(df.columns)
        global len_of_lst1
        len_of_lst1 = len(l0) + 3 
        print (l0)
        row1 = input('copy and paste the first row of the file: ')
        l1 = list(row1.split("\t"))
        print(l1) 

        dt = []
        regex_regex_float = '[+-]?[0-9]+\.[0-9]+'
        regex_for_special_string = '[@_!#$%^&*()<>?/\|}{~:]'
        regex_for_percentage = '\\d+(\\.?\\d+)?%'
        regex_for_numbers_commas_minus = '^([-+] ?)?[0-9]+(,[0-9]+)?$'

        #l0 = ['Sam', 'Video Title', 'URL', 'Platform', 'Video Views', 'Likes', 'Comments', 'Shares', 'Total Engagement (Likes + Comments + Shares) ', 'Video Eng Rate']
        #l1 = ['ssssss', 'Full 12th Gen Core family revealed at CES 2022: Mobile and Desktop | Intel Talking Tech', 'https://www.youtube.com/watch?v=YVXHv067Awo', 'YouTube', '2,105', '72', '7', '16', '95', '5%']

        dt = dict() 
        a_true = ""
        j_needed_value = ""
        for j in range(len(l0)):
            if 'date' in l0[j].lower():
                print(j,l0[j].lower()) 
                j_needed_value = j_needed_value + str(j) 
                a_true = a_true + 'True'
                a = l1[j]
                zzz = j
                # print(j)
                if len(a) == 8:

                    dt[zzz] = "Date"
                elif len(a) == 10:
                    dt[zzz] = "Date"

        if len(j_needed_value)> 0:
            j_needed_value1 = int(j_needed_value)
        else:
            j_needed_value1 = j_needed_value
        print(j_needed_value1)
        # print(int(float(j_needed_value)))
        # print(len(j_needed_value)) 
        print(type(j_needed_value1))  
        print(a_true) 
        for i in range(len(l1)):
            print(l1[i])
            if i == j_needed_value1:
                if a_true == 'True':
                    i = i + 1 

                    if l1[i].isnumeric() == True:
                        dt[i] = "Integer"                     
                    elif (re.search(regex_for_percentage,l1[i])):
                        dt[i] = "Percentage"
                    elif (re.search(regex_for_numbers_commas_minus,l1[i])):
                        dt[i] = "Integer"           
                    elif l1[i].isalpha() == True or l1[i].isalnum() == True:
                        dt[i] = "String"
                    elif (re.search(regex_for_special_string,l1[i])):
                        dt[i] = "String"
                    elif (re.search(regex_regex_float,l1[i])):
                        dt[i] = "Float"
                    else:
                        dt[i] = "Blank"

                '''---------------------------------'''

            elif l1[i].isnumeric() == True:

                dt[i] = "Integer"
            elif (re.search(regex_for_percentage,l1[i])):
                dt[i] = "Percentage"
            elif (re.search(regex_for_numbers_commas_minus,l1[i])):
                dt[i] = "Integer"           
            elif l1[i].isalpha() == True or l1[i].isalnum() == True:
                dt[i] = "String"
            elif (re.search(regex_for_special_string,l1[i])):
                dt[i] = "String"
            elif (re.search(regex_regex_float,l1[i])):
                dt[i] = "Float"
            else:
                dt[i] = "Blank"
        ct = dict(sorted(dt.items()))
        # print(dt.sort())
        a = list(dt.items())
        lst_of_sorted_items = []
        import operator
        from functools import reduce
        out = [item for t in a for item in t]
        for i in range(0,len(out)):
            # print(out[i])
            if i%2!=0 :
                lst_of_sorted_items.append(out[i])
        print(lst_of_sorted_items)
        
        
        df1=pd.DataFrame(columns=['Table',"KPI's","Datatype","Sample Data"])
        df1["KPI's"] = l0
        df1["Sample Data"] = l1
        df1["Datatype"] = lst_of_sorted_items
        
        print(df1) 
        
        data_schema_file_name_extractor = file_path.split('\\').pop()
        final_file_value = data_schema_file_name_extractor.split('.')[0]
        global final_name1
        final_name1 = f"Data Schema -{final_file_value}- Tech Community.xlsx"
        df1.loc[0,'Table'] = final_file_value

        df1.to_excel(final_name1,index=False,index_label=False)
        
        writer = pd.ExcelWriter(final_name1, engine='xlsxwriter')
        header = pd.MultiIndex.from_product([[f"Data Schema -{final_file_value}- Tech Community"],df1.columns])
        df1 = pd.DataFrame(df1.to_numpy(), index=df1.index , columns = header) 
        # df.to_excel(writer,sheet_name=f'Sheet1',index=True,index_label=True)
        final_new_name = f"Data Schema - {final_file_value} - Tech Community.xlsx"

        df1.reset_index(drop=True, inplace=True)
        
        df1.to_excel(final_new_name)

        import openpyxl
        from openpyxl.styles import PatternFill,Alignment,NamedStyle
        from openpyxl.styles.borders import Border, Side

        wb = openpyxl.load_workbook(final_new_name)
        ws = wb['Sheet1']

        fill_pattern = PatternFill(patternType='solid', fgColor='BDD7EE')

        ws['B2'].fill = fill_pattern
        ws['C2'].fill = fill_pattern
        ws['D2'].fill = fill_pattern
        ws['E2'].fill = fill_pattern
        ws["B4"].alignment = Alignment(horizontal="center",vertical='center')
        ws.merge_cells(f"B4:B{len_of_lst1}") 


        wb.save(final_new_name) 

        from openpyxl.styles import Font, Color
        from openpyxl.styles import Border,Side

        wb = openpyxl.load_workbook(final_new_name)
        ws = wb['Sheet1']
        print('1')

        top=Side(border_style='thin',color="000000")
        bottom=Side(border_style='thin', color="000000")
        right_side = Side(border_style='thin', color="000000")
        left_side = Side(border_style='thin', color="000000")
        border=Border(top=top,bottom=bottom,right=right_side,left=left_side)

        font_style1 = Font(size = 14,bold=True)
        font_style2 = Font(size = 13,bold=True)
        font_style3 = Font(size=12)
        font_style4 = Font(size=12,bold=True)


        print('2')
        for i in range(4,len_of_lst1+1):
            ws.cell(row=i, column=5).font = font_style3
            ws.cell(row=i,column=4).font = font_style3
            ws.cell(row=i, column=3).font = font_style3
            ws.cell(row=i,column=2).font = font_style4
            ws.cell(row=i, column=5).border = border
            ws.cell(row=i,column=4).border = border
            ws.cell(row=i, column=3).border = border
            ws.cell(row=i,column=2).border = border
        
        for i in range(1,6):
            ws.cell(row=2,column=i).font = font_style2
            ws.cell(row=1,column=i).font = font_style1
            ws.cell(row=2,column=i).border = border
            ws.cell(row=1,column=i).border = border

        # Imorting the necessary modules
        # stackoverflow code to automatic change the width of the columns
        try:
                from openpyxl.cell import get_column_letter
        except ImportError:
                from openpyxl.utils import get_column_letter
                from openpyxl.utils import column_index_from_string
        from openpyxl import load_workbook
        import openpyxl
        from openpyxl import Workbook



        for column_cells in ws.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length*1.23
        
        # for i in range(4,len_of_lst+1):
        #     ws.cell(row=i, column=5).border = border
        #     ws.cell(row=i,column=4).border = border
        #     ws.cell(row=i, column=3).border = border
        #     ws.cell(row=i,column=2).border = border
        # for i in range(1,6):
        #     ws.cell(row=2,column=i).border = border
        #     ws.cell(row=1,column=i).border = border
        ws.title = final_file_value
        wb.save(final_new_name)
        # import os
        # with open(filename, 'x') as f:
        # os.remove(final_name)

        
        
file_path = input('enter your file path: ')
if file_path[0] == '"' and file_path[-1] == '"':
    file_path = file_path[1:]
    file_path = file_path[:-1]

open_file(file_path) 

import os

pwd1 = os.getcwd()
print(pwd1)
file_path = pwd1+ f'\{final_name1}'
file_path1 = pwd1+ f'\{final_name1}'

if os.path.isfile(file_path):
  os.remove(file_path)
  print("File has been deleted")
elif os.path.isfile(file_path):
  os.remove(file_path)
  print("File has been deleted")
else:
  print("File does not exist") 
        


   


        

    
